from pathlib import Path
import unicodedata
import asyncio
import os
from typing import Optional

from fastapi import FastAPI, UploadFile, File, HTTPException, BackgroundTasks, Header
from fastapi.middleware.cors import CORSMiddleware

from cv_parser import read_cv_text, parse_cv_text
from jobs_data import MOCK_JOBS
from application_profile import build_application_profile
from job_filter import filter_jobs
from aggregator.engine import run_aggregation

app = FastAPI(title="Jobs AI Chile")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)


def normalize_text(text: str) -> str:
    return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("utf-8")


def normalize_skill_list(skills):
    return [normalize_text(skill).lower().strip() for skill in skills]


def get_profile_field(profile, field_name, default=None):
    value = profile.get(field_name, default)

    # Caso parsed_profile: {"value": ..., "confidence": ...}
    if isinstance(value, dict) and "value" in value:
        return value.get("value", default)

    # Caso application_profile: valor directo
    return value


def calculate_match(profile, job):
    profile_skills = set(normalize_skill_list(get_profile_field(profile, "skills", [])))
    job_skills = set(normalize_skill_list(job.get("skills", [])))

    matched_skills = sorted(profile_skills.intersection(job_skills))
    missing_skills = sorted(job_skills - profile_skills)

    score = 0
    reasons = []
    gaps = []

    # 1. Skills
    if job_skills:
        skill_score = int((len(matched_skills) / len(job_skills)) * 60)
    else:
        skill_score = 0

    score += skill_score

    if matched_skills:
        reasons.append(f"Coincides en {len(matched_skills)} skill(s) clave")

    if missing_skills:
        gaps.append(f"Te faltan skills como: {', '.join(missing_skills[:4])}")

    # 2. Seniority
    seniority_order = {
        "junior": 1,
        "semi-senior": 2,
        "senior": 3
    }

    profile_seniority = get_profile_field(profile, "seniority", "junior")
    profile_level = seniority_order.get(profile_seniority, 1)
    job_level = seniority_order.get(job.get("seniority", "junior"), 1)

    if profile_level >= job_level:
        score += 20
        reasons.append("Tu seniority es compatible")
    else:
        gaps.append("El cargo pide mayor seniority")

    # 3. Rol actual / experiencia relacionada
    current_role = normalize_text(get_profile_field(profile, "current_role", "") or "").lower()
    title = normalize_text(job.get("title", "")).lower()
    description = normalize_text(job.get("description", "")).lower()

    if current_role:
        if "inversion" in current_role and ("inversion" in title or "investment" in title):
            score += 10
            reasons.append("Tu rol actual se relaciona directamente con el cargo")
        elif "portfolio" in current_role and ("portfolio" in title or "portfolio" in description):
            score += 10
            reasons.append("Tu experiencia actual calza con el foco del cargo")
        elif "analista" in current_role or "analyst" in current_role:
            score += 5
            reasons.append("Tienes experiencia analítica relevante")

    # 4. Años de experiencia
    years = get_profile_field(profile, "years_experience", 0)
    if years >= 1:
        score += 10
        reasons.append("Tienes experiencia profesional relevante")

    score = min(score, 100)

    return {
        "job_id": job["id"],
        "job_title": job["title"],
        "company": job["company"],
        "location": job["location"],
        "modality": job["modality"],
        "match_score": score,
        "matched_skills": matched_skills,
        "missing_skills": missing_skills,
        "reasons": reasons,
        "gaps": gaps
    }


@app.get("/")
def home():
    key = BREVO_SMTP_KEY
    return {
        "status": "ok",
        "service": "AutoApply Chile Backend",
        "version": "2.0.0",
        "brevo_key_prefix": key[:12] if key else "NOT SET",
        "brevo_user": BREVO_SMTP_USER,
    }


@app.post("/upload-cv")
async def upload_cv(file: UploadFile = File(...)):
    allowed_types = {
        "application/pdf": ".pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
    }

    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Solo se permiten archivos PDF o DOCX")

    extension = allowed_types[file.content_type]
    safe_name = Path(file.filename).stem.replace(" ", "_")
    file_path = UPLOAD_DIR / f"{safe_name}{extension}"

    content = await file.read()
    file_path.write_bytes(content)

    extracted_text, extraction_method, extraction_quality = read_cv_text(file_path)
    parsed = parse_cv_text(
        extracted_text,
        extraction_quality=extraction_quality,
        extraction_method=extraction_method
    )

    return {
        "message": "CV subido correctamente",
        "filename": file_path.name,
        "path": str(file_path),
        "text_preview": extracted_text[:2500],
        "profile": parsed["profile"],
        "warnings": parsed["warnings"],
        "extraction_method": parsed["extraction_method"],
        "extraction_quality_score": parsed["extraction_quality_score"]
    }


@app.post("/parse-cv")
def parse_cv(payload: dict):
    text = payload.get("cv_text", "")
    if not text:
        raise HTTPException(status_code=400, detail="Falta cv_text")

    parsed = parse_cv_text(text)
    return parsed


@app.post("/match-jobs")
def match_jobs(profile: dict):
    results = []

    for job in MOCK_JOBS:
        match_result = calculate_match(profile, job)
        results.append(match_result)

    results.sort(key=lambda x: x["match_score"], reverse=True)

    return {
        "total_jobs": len(results),
        "matches": results
    }


@app.post("/analyze-cv")
async def analyze_cv(file: UploadFile = File(...)):
    allowed_types = {
        "application/pdf": ".pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
    }

    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Solo se permiten archivos PDF o DOCX")

    extension = allowed_types[file.content_type]
    safe_name = Path(file.filename).stem.replace(" ", "_")
    file_path = UPLOAD_DIR / f"{safe_name}{extension}"

    content = await file.read()
    file_path.write_bytes(content)

    extracted_text, extraction_method, extraction_quality = read_cv_text(file_path)
    parsed = parse_cv_text(
        extracted_text,
        extraction_quality=extraction_quality,
        extraction_method=extraction_method
    )

    results = []
    for job in MOCK_JOBS:
        results.append(calculate_match(parsed["profile"], job))

    results.sort(key=lambda x: x["match_score"], reverse=True)

    return {
        "message": "CV analizado correctamente",
        "filename": file_path.name,
        "profile": parsed["profile"],
        "warnings": parsed["warnings"],
        "extraction_method": parsed["extraction_method"],
        "extraction_quality_score": parsed["extraction_quality_score"],
        "top_matches": results[:5]
    }


@app.post("/build-application-profile")
def build_profile(payload: dict):
    parsed_profile = payload.get("parsed_profile")
    user_answers = payload.get("user_answers", {})

    if not parsed_profile:
        raise HTTPException(status_code=400, detail="Falta parsed_profile")

    final_profile = build_application_profile(parsed_profile, user_answers)

    return {
        "message": "Perfil final construido correctamente",
        "application_profile": final_profile
    }


@app.post("/search-jobs")
def search_jobs(payload: dict):
    application_profile = payload.get("application_profile")

    if not application_profile:
        raise HTTPException(status_code=400, detail="Falta application_profile")

    # 1. filtrar jobs según preferencias
    filtered_jobs = filter_jobs(MOCK_JOBS, application_profile)

    # 2. hacer matching usando professional_info
    professional_info = application_profile.get("professional_info", {})

    results = []
    for job in filtered_jobs:
        match_result = calculate_match(professional_info, job)
        results.append(match_result)

    # 3. ordenar
    results.sort(key=lambda x: x["match_score"], reverse=True)

    return {
        "total_jobs_after_filter": len(filtered_jobs),
        "results": results[:10]
    }

# ============================================================
# AGGREGATION ENDPOINTS
# ============================================================

AGGREGATE_SECRET = os.environ.get("AGGREGATE_SECRET", "autoapply-aggregate-secret")


@app.post("/aggregate")
async def aggregate_jobs(
    background_tasks: BackgroundTasks,
    payload: dict = {},
    authorization: Optional[str] = Header(None),
):
    """
    Trigger job aggregation from all sources.
    Runs in background so the request returns immediately.
    
    Body params:
      - roles: list of target roles (optional)
      - fast_mode: bool (default False) — skip scrapers for speed
      - sync: bool (default False) — wait for completion (use for testing)
    """
    # Verify secret
    expected = f"Bearer {AGGREGATE_SECRET}"
    if authorization != expected:
        raise HTTPException(status_code=401, detail="Unauthorized")

    roles = payload.get("roles", [])
    fast_mode = payload.get("fast_mode", False)
    sync = payload.get("sync", False)

    if sync:
        # Wait for result (use for testing/cron)
        result = await run_aggregation(roles=roles or None, fast_mode=fast_mode)
        return {"status": "completed", **result}

    # Run in background
    background_tasks.add_task(
        asyncio.run,
        run_aggregation(roles=roles or None, fast_mode=fast_mode)
    )
    return {"status": "started", "message": "Aggregation running in background"}


@app.get("/aggregate/status")
async def aggregate_status():
    """Return counts from the jobs table for monitoring."""
    from aggregator.storage import get_client
    supabase = get_client()
    
    try:
        total = supabase.table("jobs").select("id", count="exact").execute()
        by_source = supabase.table("jobs").select("source").execute()
        
        source_counts = {}
        for row in (by_source.data or []):
            src = row["source"]
            source_counts[src] = source_counts.get(src, 0) + 1
        
        return {
            "total_jobs": total.count,
            "by_source": source_counts,
        }
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# AUTO-APPLY ENDPOINT (Email via Resend)
# ============================================================

RESEND_API_KEY = os.environ.get("RESEND_API_KEY", "")

# Brevo SMTP (primary email provider)
BREVO_SMTP_HOST = os.environ.get("BREVO_SMTP_HOST", "smtp-relay.brevo.com")
BREVO_SMTP_PORT = int(os.environ.get("BREVO_SMTP_PORT", "587"))
BREVO_SMTP_USER = os.environ.get("BREVO_SMTP_USER", "")
BREVO_SMTP_KEY = os.environ.get("BREVO_SMTP_KEY", "").strip()
BREVO_FROM_EMAIL = os.environ.get("BREVO_FROM_EMAIL", "AutoApply Chile <noreply@autoapply.cl>")


def detect_ats(apply_link: str) -> str:
    url = apply_link.lower()
    if "greenhouse.io" in url or "job-boards.greenhouse" in url:
        return "greenhouse"
    if "lever.co" in url:
        return "lever"
    if "myworkdayjobs.com" in url:
        return "workday"
    if "smartrecruiters.com" in url:
        return "smartrecruiters"
    return "generic"


def extract_recruiter_email(job: dict) -> str | None:
    """Return apply_email if present in job dict."""
    return job.get("apply_email") or job.get("contact_email") or None


async def send_application_email(
    to_email: str,
    candidate_name: str,
    candidate_email: str,
    candidate_phone: str,
    job_title: str,
    company: str,
    cover_letter: str,
    cv_base64: str,
    cv_filename: str,
) -> dict:
    import httpx

    if not BREVO_SMTP_KEY:
        return {"success": False, "message": "Brevo API key no configurada", "method": "email"}

    subject = f"Postulación: {job_title} — {candidate_name}"

    html_body = f"""
<div style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;max-width:560px;margin:0 auto;">
  <div style="background:#2563eb;padding:24px 32px;border-radius:12px 12px 0 0;">
    <h2 style="color:white;margin:0;font-size:18px;">Postulación: {job_title}</h2>
    <p style="color:#bfdbfe;margin:4px 0 0;font-size:14px;">{company}</p>
  </div>
  <div style="background:white;padding:24px 32px;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 12px 12px;">
    <table style="font-size:14px;color:#374151;margin-bottom:16px;">
      <tr><td style="padding:4px 12px 4px 0;color:#6b7280;">Candidato</td><td><strong>{candidate_name}</strong></td></tr>
      <tr><td style="padding:4px 12px 4px 0;color:#6b7280;">Email</td><td>{candidate_email}</td></tr>
      <tr><td style="padding:4px 12px 4px 0;color:#6b7280;">Teléfono</td><td>{candidate_phone or 'No indicado'}</td></tr>
    </table>
    <hr style="border:none;border-top:1px solid #f3f4f6;margin:16px 0;">
    <h3 style="font-size:15px;color:#111827;margin:0 0 8px;">Carta de presentación</h3>
    <p style="font-size:14px;color:#374151;white-space:pre-line;line-height:1.6;">{cover_letter}</p>
    <hr style="border:none;border-top:1px solid #f3f4f6;margin:16px 0;">
    <p style="font-size:12px;color:#9ca3af;text-align:center;">Postulación enviada vía AutoApply Chile</p>
  </div>
</div>
"""

    payload: dict = {
        "from": "AutoApply Chile <noreply@autoapplychile.com>",
        "to": [to_email],
        "cc": [candidate_email],
        "reply_to": candidate_email,
        "subject": subject,
        "html": html_body,
    }

    if cv_base64:
        payload["attachments"] = [{"filename": cv_filename, "content": cv_base64}]

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                "https://api.resend.com/emails",
                headers={
                    "Authorization": f"Bearer {RESEND_API_KEY}",
                    "Content-Type": "application/json",
                },
                json=payload,
            )

        if resp.status_code in (200, 201):
            return {"success": True, "message": f"Email enviado a {to_email}", "method": "email"}
        else:
            return {"success": False, "message": f"Error Resend: {resp.text}", "method": "email"}

    except Exception as e:
        return {"success": False, "message": f"Error enviando email: {str(e)}", "method": "email"}


@app.post("/auto-apply")
async def auto_apply_endpoint(payload: dict, authorization: Optional[str] = Header(None)):
    """
    Email-first auto-apply strategy:
    1. Check job.apply_email
    2. Try Hunter.io to find recruiter email for the company
    3. Fall back to common HR email pattern
    4. If nothing found, return cover letter for manual apply

    Body:
      - job: {id, title, company, apply_link, apply_email, description}
      - parsed_cv: {name, email, phone, ...}
      - cv_base64: base64-encoded PDF
      - cover_letter: generated cover letter text
      - user_preferences: {}
      - existing_answers: {}
    """
    expected = f"Bearer {AGGREGATE_SECRET}"
    if authorization != expected:
        raise HTTPException(status_code=401, detail="Unauthorized")

    job = payload.get("job", {})
    parsed_cv = payload.get("parsed_cv", {})
    cv_base64 = payload.get("cv_base64", "")
    cover_letter = payload.get("cover_letter", "")
    user_preferences = payload.get("user_preferences", {})
    existing_answers = payload.get("existing_answers", {})

    apply_link = job.get("apply_link", "") or ""
    candidate_name = parsed_cv.get("name") or ""
    candidate_email = parsed_cv.get("email") or ""
    candidate_phone = parsed_cv.get("phone") or ""

    if not candidate_email:
        raise HTTPException(status_code=400, detail="El perfil no tiene email")

    # 1. Use apply_email from job if available
    recruiter_email = job.get("apply_email") or job.get("contact_email") or None

    # 2. Try Hunter.io + common patterns
    if not recruiter_email and RESEND_API_KEY:
        from aggregator.recruiter_finder import find_recruiter_email
        recruiter_email = await find_recruiter_email(
            company=job.get("company", ""),
            apply_link=apply_link,
        )

    # 3. Send email if we have a destination
    if recruiter_email and RESEND_API_KEY:
        cv_filename = f"CV_{candidate_name.replace(' ', '_') or 'Candidato'}.pdf"
        result = await send_application_email(
            to_email=recruiter_email,
            candidate_name=candidate_name,
            candidate_email=candidate_email,
            candidate_phone=candidate_phone,
            job_title=job.get("title", ""),
            company=job.get("company", ""),
            cover_letter=cover_letter,
            cv_base64=cv_base64,
            cv_filename=cv_filename,
        )
        return {
            **result,
            "method": "email",
            "recruiter_email": recruiter_email,
            "cover_letter": cover_letter,
            "apply_link": apply_link,
        }

    # 4. No email found — return cover letter for manual apply
    return {
        "success": False,
        "method": "manual",
        "message": "No encontramos email del reclutador. Usa el link para postular con tu carta lista.",
        "cover_letter": cover_letter,
        "apply_link": apply_link,
    }


@app.post("/import-job")
async def import_job_endpoint(payload: dict, authorization: Optional[str] = Header(None)):
    """
    Import a job from a URL (LinkedIn, portal empresa, etc).
    Scrapes the page, extracts job info, saves to Supabase.

    Body: { "url": "https://linkedin.com/jobs/view/..." }
    """
    expected = f"Bearer {AGGREGATE_SECRET}"
    if authorization != expected:
        raise HTTPException(status_code=401, detail="Unauthorized")

    url = payload.get("url", "").strip()
    if not url:
        raise HTTPException(status_code=400, detail="url requerido")

    import hashlib
    import httpx
    import re

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "es-CL,es;q=0.9,en;q=0.8",
    }

    try:
        async with httpx.AsyncClient(headers=headers, follow_redirects=True, timeout=20) as client:
            resp = await client.get(url)
            html = resp.text
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"No se pudo acceder a la URL: {e}")

    # Extract JSON-LD JobPosting
    job_data = {}
    jsonld_matches = re.findall(
        r'<script[^>]+type="application/ld\+json"[^>]*>(.*?)</script>',
        html, re.DOTALL
    )
    for raw in jsonld_matches:
        try:
            import json
            obj = json.loads(raw)
            if obj.get("@type") == "JobPosting":
                job_data = obj
                break
        except Exception:
            continue

    # Extract from meta tags if no JSON-LD
    def meta(name: str) -> str:
        m = re.search(
            rf'<meta[^>]+(?:name|property)=["\'](?:og:)?{name}["\'][^>]+content=["\']([^"\']+)["\']',
            html, re.IGNORECASE
        )
        return m.group(1).strip() if m else ""

    title = (
        job_data.get("title") or
        meta("title") or
        re.search(r'<title>([^<|–-]+)', html, re.IGNORECASE) and
        re.search(r'<title>([^<|–-]+)', html, re.IGNORECASE).group(1).strip() or
        "Sin título"
    )

    company_raw = job_data.get("hiringOrganization") or {}
    company = (
        company_raw.get("name") if isinstance(company_raw, dict) else str(company_raw)
    ) or meta("og:site_name") or "Empresa confidencial"

    location_raw = job_data.get("jobLocation") or {}
    if isinstance(location_raw, dict):
        addr = location_raw.get("address") or {}
        location = addr.get("addressLocality") or addr.get("addressRegion") or "Chile"
    else:
        location = "Chile"

    description = job_data.get("description") or meta("description") or ""
    # Strip HTML tags from description
    description = re.sub(r'<[^>]+>', ' ', description).strip()

    from aggregator.greenhouse import extract_skills_from_content, detect_seniority
    from aggregator.storage import get_client

    job_id = hashlib.md5(url.encode()).hexdigest()[:12]
    external_id = f"manual_{job_id}"

    apply_email = payload.get("apply_email", "").strip() or None

    from aggregator.base import NormalizedJob
    job = NormalizedJob(
        external_id=external_id,
        source="manual",
        title=title[:200],
        company=company[:200],
        location=location,
        country="CL",
        description=description[:3000],
        apply_link=url,
        apply_email=apply_email,
        skills=extract_skills_from_content(description),
        seniority=detect_seniority(title, description),
    )

    # Save to Supabase
    from aggregator.storage import upsert_jobs
    inserted, updated = upsert_jobs([job])

    return {
        "success": True,
        "job": {
            "external_id": external_id,
            "title": title,
            "company": company,
            "location": location,
            "apply_link": url,
            "skills": job.skills,
        },
        "inserted": inserted,
        "updated": updated,
    }


@app.post("/import-excel")
async def import_excel_endpoint(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(None),
):
    """
    Import jobs from an Excel (.xlsx) or CSV file.
    Expected columns: empresa, cargo, ciudad, link, descripcion, modalidad (optional)
    """
    expected = f"Bearer {AGGREGATE_SECRET}"
    if authorization != expected:
        raise HTTPException(status_code=401, detail="Unauthorized")

    import hashlib
    import io
    import csv

    content = await file.read()
    filename = file.filename or ""
    jobs_to_insert = []

    try:
        if filename.endswith(".csv"):
            text = content.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error leyendo archivo: {e}")

    # Column name aliases
    def get_col(row: dict, *names: str) -> str:
        for n in names:
            v = row.get(n, "").strip()
            if v and v.lower() != "none":
                return v
        return ""

    from aggregator.greenhouse import extract_skills_from_content, detect_seniority
    from aggregator.base import NormalizedJob
    from aggregator.storage import upsert_jobs

    for row in rows:
        company = get_col(row, "empresa", "company", "compañia", "compania")
        title = get_col(row, "cargo", "title", "puesto", "posicion", "posición")
        if not company or not title:
            continue

        city = get_col(row, "ciudad", "location", "ciudad/region", "región", "region") or "Chile"
        link = get_col(row, "link", "url", "apply_link", "link_postulacion")
        description = get_col(row, "descripcion", "descripción", "description", "detalle")
        apply_email = get_col(row, "email_reclutador", "email", "recruiter_email", "apply_email", "contacto")
        modality_raw = get_col(row, "modalidad", "modality", "modalidad de trabajo").lower()

        modality = None
        if "remoto" in modality_raw or "remote" in modality_raw:
            modality = "remote"
        elif "híbrido" in modality_raw or "hibrido" in modality_raw or "hybrid" in modality_raw:
            modality = "hybrid"
        elif "presencial" in modality_raw or "onsite" in modality_raw:
            modality = "presencial"

        job_id = hashlib.md5(f"{company}{title}{link}".encode()).hexdigest()[:12]

        job = NormalizedJob(
            external_id=f"excel_{job_id}",
            source="excel",
            title=title[:200],
            company=company[:200],
            location=city,
            country="CL",
            description=description[:3000],
            apply_link=link or None,
            apply_email=apply_email or None,
            modality=modality,
            skills=extract_skills_from_content(description),
            seniority=detect_seniority(title, description),
        )
        jobs_to_insert.append(job)

    if not jobs_to_insert:
        raise HTTPException(status_code=400, detail="No se encontraron filas válidas. Verifica las columnas: empresa, cargo, ciudad, link, descripcion")

    inserted, updated = upsert_jobs(jobs_to_insert)
    return {
        "success": True,
        "total_rows": len(rows),
        "inserted": inserted,
        "updated": updated,
        "jobs_processed": len(jobs_to_insert),
    }


@app.post("/generate-answers")
async def generate_answers_endpoint(payload: dict, authorization: Optional[str] = Header(None)):
    """
    Generate all standard answers from CV + AI (without applying).
    Use this to preview what answers will be used.
    """
    expected = f"Bearer {AGGREGATE_SECRET}"
    if authorization != expected:
        raise HTTPException(status_code=401, detail="Unauthorized")

    from aggregator.answer_generator import ai_fill_answers

    answers = ai_fill_answers(
        parsed_cv=payload.get("parsed_cv", {}),
        user_preferences=payload.get("user_preferences", {}),
        job=payload.get("job", {}),
        existing_answers=payload.get("existing_answers", {}),
    )
    return {"answers": answers}
